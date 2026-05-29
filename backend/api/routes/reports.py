"""Reports API routes."""
from typing import Optional, List, Dict, Any
from datetime import date, datetime
from fastapi import APIRouter, Query, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import io

from backend.database import get_db
from backend.api.auth import get_current_user, UserAuth
from backend.database.user_repository import UserRepository
from backend.core.date_utils import to_vn_date_str

router = APIRouter(prefix="/reports", tags=["reports"])


class InvoiceFlowReportItem(BaseModel):
    """Individual invoice item for the flow report."""
    id: str
    date: str
    type: str  # 'in' for purchase, 'out' for sale
    company_name: str
    tax_code: str
    amount: float  # Positive for incoming, negative for outgoing
    tax_amount: float  # Positive for incoming, negative for outgoing
    invoice_number: str
    invoice_symbol: str


class InvoiceFlowReportResponse(BaseModel):
    """Response model for the invoice flow report."""
    items: List[InvoiceFlowReportItem]
    total_incoming_amount: float
    total_outgoing_amount: float
    total_incoming_tax: float
    total_outgoing_tax: float
    net_tax_obligation: float  # outgoing_tax - incoming_tax


class VATTaxTimelineItem(BaseModel):
    """Individual VAT tax timeline item."""
    date: str
    incoming_tax: float
    outgoing_tax: float
    net_tax: float  # outgoing_tax - incoming_tax


class VATTaxTimelineResponse(BaseModel):
    """Response model for the VAT tax timeline report."""
    items: List[VATTaxTimelineItem]
    total_incoming_tax: float
    total_outgoing_tax: float
    net_tax_obligation: float


class InvoiceRelationPairItem(BaseModel):
    """One relation pair: adjusted/replaced invoice and original invoice."""
    relation_type: str
    relation_type_label: str
    actor_invoice_id: str
    actor_invoice_number: Optional[int] = None
    actor_invoice_symbol: Optional[str] = None
    actor_invoice_date: Optional[str] = None
    actor_invoice_status: Optional[int] = None
    base_invoice_id: Optional[str] = None
    base_invoice_number: Optional[int] = None
    base_invoice_symbol: Optional[str] = None
    base_invoice_date: Optional[str] = None
    base_invoice_status: Optional[int] = None
    seller_tax_code: Optional[str] = None
    buyer_tax_code: Optional[str] = None
    amount_before_tax: float = 0
    tax_amount: float = 0
    total_amount: float = 0


class InvoiceRelationPairsResponse(BaseModel):
    """Response for adjustment/replacement pairs report."""
    items: List[InvoiceRelationPairItem]
    total_pairs: int
    total_adjustment_pairs: int
    total_replacement_pairs: int


def build_company_restriction_clause(user_id: int, role: str, conn, tax_code_field: str, buyer_tax_code_field: str):
    """Build company restriction clause for non-admin users."""
    if role == "admin":
        return "1=1", []
    
    user_repo = UserRepository(conn)
    user_companies = user_repo.get_user_companies(user_id)
    company_tax_codes = [comp['tax_code'] for comp in user_companies]
    
    if not company_tax_codes:
        return "1=0", []  # No access
    
    company_placeholders = ",".join(["%s"] * len(company_tax_codes))
    clause = f"({tax_code_field} IN ({company_placeholders}) OR {buyer_tax_code_field} IN ({company_placeholders}))"
    params = company_tax_codes + company_tax_codes
    
    return clause, params


def _fetch_relation_pairs(
    conn,
    current_user: UserAuth,
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by company tax code"),
):
    if not from_date or not to_date:
        raise HTTPException(status_code=400, detail="from_date and to_date are required")
    if from_date > to_date:
        raise HTTPException(status_code=400, detail="from_date must be <= to_date")

    date_from = from_date.isoformat()
    date_to = to_date.isoformat() + "T23:59:59"

    conditions = []
    params = []

    conditions.append("(COALESCE(d.nky, d.tdlap) BETWEEN %s AND %s)")
    params.extend([date_from, date_to])

    company_clause, company_params = build_company_restriction_clause(
        current_user.id, current_user.role, conn, "d.nbmst", "d.nmmst"
    )
    conditions.append(company_clause)
    params.extend(company_params)

    if tax_code:
        conditions.append("(d.nbmst = %s OR d.nmmst = %s)")
        params.extend([tax_code, tax_code])

    where_clause = " AND ".join(conditions)

    query = f"""
        SELECT
            CASE
                WHEN d.tthai = 3 THEN 'adjustment'
                WHEN d.tthai = 2 THEN 'replacement'
                ELSE 'related'
            END AS relation_type,
            d.id AS actor_invoice_id,
            d.shdon AS actor_invoice_number,
            d.khhdon AS actor_invoice_symbol,
            COALESCE(d.nky, d.tdlap) AS actor_invoice_date,
            d.tthai AS actor_invoice_status,
            b.id AS base_invoice_id,
            b.shdon AS base_invoice_number,
            b.khhdon AS base_invoice_symbol,
            COALESCE(b.nky, b.tdlap) AS base_invoice_date,
            b.tthai AS base_invoice_status,
            d.nbmst AS seller_tax_code,
            d.nmmst AS buyer_tax_code,
            COALESCE(d.tgtcthue, 0) AS amount_before_tax,
            COALESCE(d.tgtthue, 0) AS tax_amount,
            COALESCE(d.tgtttbso, 0) AS total_amount
        FROM invoices d
        LEFT JOIN invoices b
            ON b.nbmst = d.nbmst
            AND b.khhdon = d.khhdgoc
            AND b.khmshdon::text = d.khmshdgoc
            AND b.shdon::text = d.shdgoc::text
        WHERE {where_clause}
            AND d.shdgoc IS NOT NULL
            AND d.tthai IN (2, 3)
        ORDER BY COALESCE(d.nky, d.tdlap) DESC, d.shdon DESC
    """

    with conn.cursor() as cur:
        cur.execute(query, params)
        rows = cur.fetchall()

    items: List[InvoiceRelationPairItem] = []
    for row in rows:
        r = dict(row)
        relation_type = r.get("relation_type")
        relation_type_label = "Điều chỉnh / Bị điều chỉnh" if relation_type == "adjustment" else "Thay thế / Bị thay thế"
        items.append(
            InvoiceRelationPairItem(
                relation_type=relation_type,
                relation_type_label=relation_type_label,
                actor_invoice_id=r.get("actor_invoice_id"),
                actor_invoice_number=r.get("actor_invoice_number"),
                actor_invoice_symbol=r.get("actor_invoice_symbol"),
                actor_invoice_date=to_vn_date_str(r.get("actor_invoice_date")),
                actor_invoice_status=r.get("actor_invoice_status"),
                base_invoice_id=r.get("base_invoice_id"),
                base_invoice_number=r.get("base_invoice_number"),
                base_invoice_symbol=r.get("base_invoice_symbol"),
                base_invoice_date=(to_vn_date_str(r.get("base_invoice_date")) if r.get("base_invoice_date") else None),
                base_invoice_status=r.get("base_invoice_status"),
                seller_tax_code=r.get("seller_tax_code"),
                buyer_tax_code=r.get("buyer_tax_code"),
                amount_before_tax=float(r.get("amount_before_tax") or 0),
                tax_amount=float(r.get("tax_amount") or 0),
                total_amount=float(r.get("total_amount") or 0),
            )
        )

    total_adjustment_pairs = sum(1 for i in items if i.relation_type == "adjustment")
    total_replacement_pairs = sum(1 for i in items if i.relation_type == "replacement")

    return items, total_adjustment_pairs, total_replacement_pairs


@router.get("/invoice-relation-pairs", response_model=InvoiceRelationPairsResponse)
async def get_invoice_relation_pairs_report(
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by company tax code"),
    current_user: UserAuth = Depends(get_current_user),
    conn=Depends(get_db),
):
    """
    Get invoice relation pairs:
    - adjustment / adjusted
    - replacement / replaced
    """
    items, total_adjustment_pairs, total_replacement_pairs = _fetch_relation_pairs(
        conn=conn,
        current_user=current_user,
        from_date=from_date,
        to_date=to_date,
        tax_code=tax_code,
    )

    return InvoiceRelationPairsResponse(
        items=items,
        total_pairs=len(items),
        total_adjustment_pairs=total_adjustment_pairs,
        total_replacement_pairs=total_replacement_pairs,
    )


@router.get("/invoice-relation-pairs/export")
async def export_invoice_relation_pairs_excel(
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by company tax code"),
    current_user: UserAuth = Depends(get_current_user),
    conn=Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    items, _, _ = _fetch_relation_pairs(
        conn=conn,
        current_user=current_user,
        from_date=from_date,
        to_date=to_date,
        tax_code=tax_code,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cap hoa don"

    headers = [
        "Loai quan he",
        "So HD dieu chinh/thay the",
        "Ngay HD dieu chinh/thay the",
        "So HD bi dieu chinh/bi thay the",
        "Ngay HD bi dieu chinh/bi thay the",
        "MST nguoi ban",
        "MST nguoi mua",
        "Tien truoc thue",
        "Tien thue",
        "Tong thanh toan",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin

    for idx, item in enumerate(items, start=2):
        ws.cell(row=idx, column=1, value=item.relation_type_label).border = thin
        ws.cell(row=idx, column=2, value=f"{item.actor_invoice_symbol or ''}-{item.actor_invoice_number or ''}").border = thin
        ws.cell(row=idx, column=3, value=item.actor_invoice_date or "").border = thin
        ws.cell(row=idx, column=4, value=f"{item.base_invoice_symbol or ''}-{item.base_invoice_number or ''}").border = thin
        ws.cell(row=idx, column=5, value=item.base_invoice_date or "").border = thin
        ws.cell(row=idx, column=6, value=item.seller_tax_code or "").border = thin
        ws.cell(row=idx, column=7, value=item.buyer_tax_code or "").border = thin
        ws.cell(row=idx, column=8, value=item.amount_before_tax).border = thin
        ws.cell(row=idx, column=9, value=item.tax_amount).border = thin
        ws.cell(row=idx, column=10, value=item.total_amount).border = thin

    widths = [24, 26, 20, 30, 24, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"bao_cao_cap_hoa_don_{from_date or 'all'}_{to_date or 'all'}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/invoice-flow", response_model=InvoiceFlowReportResponse)
async def get_invoice_flow_report(
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by company tax code"),
    current_user: UserAuth = Depends(get_current_user),
    conn = Depends(get_db),
):
    """
    Get invoice flow report combining both incoming and outgoing invoices.
    Incoming invoices (purchases) have positive amounts, outgoing invoices (sales) have negative amounts.
    """
    # Build base query conditions
    conditions = []
    params = []

    if from_date:
        conditions.append("tdlap >= %s")
        params.append(from_date.isoformat())
    if to_date:
        conditions.append("tdlap <= %s")
        params.append(to_date.isoformat() + "T23:59:59")
    
    # For non-admin users, restrict access to assigned companies
    company_clause, company_params = build_company_restriction_clause(
        current_user.id, current_user.role, conn, "nbmst", "nmmst"
    )
    conditions.append(company_clause)
    params.extend(company_params)
    
    # If specific tax code is provided, add it to the condition
    if tax_code:
        conditions.append("(nbmst = %s OR nmmst = %s)")
        params.extend([tax_code, tax_code])
    
    where_clause = " AND ".join(conditions) if conditions else "1=1"
    
    # Query to get all invoices
    query = f"""
        SELECT
            id, tdlap, nbmst, nbten, nmmst, nmten, shdon, khhdon,
            tgtcthue, tgtthue
        FROM invoices
        WHERE {where_clause}
        ORDER BY tdlap ASC, shdon ASC
    """

    with conn.cursor() as cur:
        cur.execute(query, params)
        rows = cur.fetchall()
    
    # Process results
    items = []
    total_incoming_amount = 0
    total_outgoing_amount = 0
    total_incoming_tax = 0
    total_outgoing_tax = 0
    
    # Determine target tax codes for classifying invoices
    # Use the selected company's tax_code; fall back to user's assigned companies
    if tax_code:
        target_tax_codes = {tax_code}
    else:
        user_repo = UserRepository(conn)
        user_companies = user_repo.get_user_companies(current_user.id)
        target_tax_codes = {comp['tax_code'] for comp in user_companies}
    
    for row in rows:
        row_dict = dict(row)
        
        # Đầu ra (Bán): nbmst (MST người bán) = MST công ty
        # Đầu vào (Mua): nmmst (MST người mua) = MST công ty
        invoice_type = 'out' if row_dict['nbmst'] in target_tax_codes else 'in'
        
        # Calculate amounts based on invoice type
        amount = float(row_dict['tgtcthue'] or 0)
        tax_amount = float(row_dict['tgtthue'] or 0)
        
        if invoice_type == 'out':
            # For outgoing invoices, amounts are negative
            amount = -amount
            tax_amount = -tax_amount
        
        item = InvoiceFlowReportItem(
            id=row_dict['id'],
            date=to_vn_date_str(row_dict.get('tdlap')),
            type=invoice_type,
            company_name=row_dict['nmten'] if invoice_type == 'in' else row_dict['nbten'],
            tax_code=row_dict['nmmst'] if invoice_type == 'in' else row_dict['nbmst'],
            amount=amount,
            tax_amount=tax_amount,
            invoice_number=str(row_dict['shdon']) if row_dict['shdon'] else '',
            invoice_symbol=row_dict['khhdon'] or ''
        )
        
        items.append(item)
        
        # Update totals
        if invoice_type == 'in':
            total_incoming_amount += abs(amount)  # Using abs because incoming amounts are positive
            total_incoming_tax += abs(tax_amount)
        else:
            total_outgoing_amount += abs(amount)  # Using abs because outgoing amounts are stored as negative
            total_outgoing_tax += abs(tax_amount)
    
    # Calculate net tax obligation (what the company owes = outgoing tax - incoming tax)
    net_tax_obligation = total_outgoing_tax - total_incoming_tax
    
    return InvoiceFlowReportResponse(
        items=items,
        total_incoming_amount=total_incoming_amount,
        total_outgoing_amount=total_outgoing_amount,
        total_incoming_tax=total_incoming_tax,
        total_outgoing_tax=total_outgoing_tax,
        net_tax_obligation=net_tax_obligation
    )


@router.get("/vat-timeline", response_model=VATTaxTimelineResponse)
async def get_vat_timeline_report(
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by company tax code"),
    current_user: UserAuth = Depends(get_current_user),
    conn = Depends(get_db),
):
    """
    Get VAT tax timeline showing incoming vs outgoing taxes over time.
    """
    # Build base query conditions
    conditions = []
    params = []

    if from_date:
        conditions.append("tdlap >= %s")
        params.append(from_date.isoformat())
    if to_date:
        conditions.append("tdlap <= %s")
        params.append(to_date.isoformat() + "T23:59:59")
    
    # For non-admin users, restrict access to assigned companies
    company_clause, company_params = build_company_restriction_clause(
        current_user.id, current_user.role, conn, "nbmst", "nmmst"
    )
    conditions.append(company_clause)
    params.extend(company_params)
    
    # If specific tax code is provided, add it to the condition
    if tax_code:
        conditions.append("(nbmst = %s OR nmmst = %s)")
        params.extend([tax_code, tax_code])
    
    where_clause = " AND ".join(conditions) if conditions else "1=1"
    
    # Simpler approach - get all invoices and classify them in Python
    query = f"""
        SELECT
            i.tdlap, i.nbmst, i.nmmst, i.tgtthue
        FROM invoices i
        WHERE {where_clause}
        ORDER BY i.tdlap ASC
    """
    
    with conn.cursor() as cur:
        cur.execute(query, params)
        rows = cur.fetchall()
    
    # Determine target tax codes for classifying invoices
    if tax_code:
        target_tax_codes = {tax_code}
    else:
        user_repo = UserRepository(conn)
        user_companies = user_repo.get_user_companies(current_user.id)
        target_tax_codes = {comp['tax_code'] for comp in user_companies}
    
    # Group by date and calculate incoming vs outgoing taxes
    date_data: Dict[str, Dict[str, float]] = {}
    
    for row in rows:
        row_dict = dict(row)
        date_str = to_vn_date_str(row_dict.get('tdlap')) or None
        if not date_str:
            continue
            
        tax_amount = float(row_dict['tgtthue'] or 0)
        
        if date_str not in date_data:
            date_data[date_str] = {'incoming_tax': 0, 'outgoing_tax': 0}
        
        # Đầu ra (Bán): nbmst (MST người bán) = MST công ty
        if row_dict['nbmst'] in target_tax_codes:
            date_data[date_str]['outgoing_tax'] += tax_amount
        # Đầu vào (Mua): nmmst (MST người mua) = MST công ty
        elif row_dict['nmmst'] in target_tax_codes:
            date_data[date_str]['incoming_tax'] += tax_amount
    
    # Convert to response format
    items = []
    total_incoming_tax = 0
    total_outgoing_tax = 0
    
    for date_str, data in date_data.items():
        incoming_tax = data['incoming_tax']
        outgoing_tax = data['outgoing_tax']
        net_tax = outgoing_tax - incoming_tax  # Amount owed (positive) or credit (negative)
        
        item = VATTaxTimelineItem(
            date=date_str,
            incoming_tax=incoming_tax,
            outgoing_tax=outgoing_tax,
            net_tax=net_tax
        )
        
        items.append(item)
        total_incoming_tax += incoming_tax
        total_outgoing_tax += outgoing_tax
    
    net_tax_obligation = total_outgoing_tax - total_incoming_tax
    
    return VATTaxTimelineResponse(
        items=items,
        total_incoming_tax=total_incoming_tax,
        total_outgoing_tax=total_outgoing_tax,
        net_tax_obligation=net_tax_obligation
    )
