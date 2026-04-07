"""Invoice API routes."""
from typing import Optional
from datetime import date
import io
from fastapi import APIRouter, Query, HTTPException, Depends
from fastapi.responses import StreamingResponse

from backend.database import get_db
from backend.api.schemas import (
    InvoiceSummary,
    InvoiceDetail,
    InvoiceItemResponse,
    InvoiceListResponse,
    InvoiceStatsResponse,
)
from backend.api.auth import get_current_user, UserAuth
from backend.database.user_repository import UserRepository
from backend.database.company_repository import CompanyRepository

router = APIRouter(prefix="/invoices", tags=["invoices"])


def build_invoice_where_clause(from_date, to_date, tax_code, buyer_tax_code, search):
    """Build WHERE clause and params for invoice queries.
    
    Date range filtering uses nky (ngày ký - signing date) with fallback to
    tdlap (ngày lập) when nky is NULL: COALESCE(nky, tdlap).
    Both columns are TEXT stored as ISO8601 strings (YYYY-MM-DD...).
    """
    conditions = []
    params = []
    
    if from_date:
        conditions.append("COALESCE(nky, tdlap) >= %s")
        params.append(from_date.isoformat())
        
    if to_date:
        conditions.append("COALESCE(nky, tdlap) <= %s")
        params.append(to_date.isoformat() + "T23:59:59")
        
    if tax_code:
        conditions.append("nbmst = %s")
        params.append(tax_code)
        
    if buyer_tax_code:
        conditions.append("nmmst = %s")
        params.append(buyer_tax_code)
        
    if search:
        conditions.append("(nbten ILIKE %s OR nmten ILIKE %s OR CAST(shdon AS TEXT) LIKE %s)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
    
    where_clause = " AND ".join(conditions) if conditions else "1=1"
    return where_clause, params


@router.get("", response_model=InvoiceListResponse)
async def list_invoices(
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(20, ge=1, le=100, description="Page size"),
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by seller tax code (nbmst)"),
    buyer_tax_code: Optional[str] = Query(None, description="Filter by buyer tax code (nmmst)"),
    search: Optional[str] = Query(None, description="Search in invoice number or company name"),
    current_user: UserAuth = Depends(get_current_user),
    conn = Depends(get_db),
):
    """
    List invoices with pagination and filters.
    Enforces user-company access control for non-admin users.
    """
    offset = (page - 1) * size

    # For non-admin users, restrict access to assigned companies
    user_repo = UserRepository(conn)
    company_repo = CompanyRepository(conn)
    
    if current_user.role != "admin":
        # Get user's assigned companies
        user_companies = user_repo.get_user_companies(current_user.id)
        company_ids = [comp['id'] for comp in user_companies]
        
        if not company_ids:
            # User has no assigned companies, return empty result
            return InvoiceListResponse(
                items=[],
                total=0,
                page=page,
                size=size,
                pages=0,
            )
        
        # Get tax codes for these companies
        company_tax_codes = []
        for comp in user_companies:
            company_tax_codes.append(comp['tax_code'])
        
        # Modify the where clause to restrict to user's companies
        where_clause, params = build_invoice_where_clause(from_date, to_date, tax_code, buyer_tax_code, search)
        
        # Add company restriction for non-admin users
        if company_tax_codes:
            # Create placeholders for company tax codes
            company_placeholders = ",".join(["%s"] * len(company_tax_codes))
            
            # Modify where clause to include company restrictions
            if where_clause == "1=1":
                # If no other filters, just add company filter
                where_clause = f"(nbmst IN ({company_placeholders}) OR nmmst IN ({company_placeholders}))"
                params = company_tax_codes + company_tax_codes
            else:
                # If there are other filters, add company restriction to existing where clause
                where_clause = f"({where_clause}) AND (nbmst IN ({company_placeholders}) OR nmmst IN ({company_placeholders}))"
                params = params + company_tax_codes + company_tax_codes
        else:
            # User has no companies assigned, return empty result
            return InvoiceListResponse(
                items=[],
                total=0,
                page=page,
                size=size,
                pages=0,
            )
    else:
        # Admin users can see all invoices
        where_clause, params = build_invoice_where_clause(from_date, to_date, tax_code, buyer_tax_code, search)

    # Count total
    count_sql = f"SELECT COUNT(*) as count FROM invoices WHERE {where_clause}"

    # Fetch data
    data_sql = f"""
        SELECT id, nbmst, nbten, nmmst, nmten, shdon, khhdon, khmshdon,
               tdlap, nky, tgtcthue, tgtthue, tgtttbso, tthai
        FROM invoices
        WHERE {where_clause}
        ORDER BY nky DESC NULLS LAST, tdlap DESC NULLS LAST, shdon DESC NULLS LAST
        LIMIT %s OFFSET %s
    """

    with conn.cursor() as cur:
        # Get total count
        cur.execute(count_sql, params if params else None)
        count_row = cur.fetchone()
        total = count_row['count'] if isinstance(count_row, dict) else count_row[0]

        # Get data with pagination
        data_params = params + [size, offset]
        cur.execute(data_sql, data_params)
        rows = cur.fetchall()

    # Rows are already dicts from RealDictCursor
    items = [InvoiceSummary(**dict(row)) for row in rows]
    pages = (total + size - 1) // size if total > 0 else 1

    return InvoiceListResponse(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages,
    )


@router.get("/export")
async def export_invoices_excel(
    from_date: Optional[date] = Query(None, description="From date (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="To date (YYYY-MM-DD)"),
    tax_code: Optional[str] = Query(None, description="Filter by seller tax code (nbmst)"),
    buyer_tax_code: Optional[str] = Query(None, description="Filter by buyer tax code (nmmst)"),
    search: Optional[str] = Query(None, description="Search in invoice number or company name"),
    current_user: UserAuth = Depends(get_current_user),
    conn = Depends(get_db),
):
    """
    Export invoices to Excel file with detailed line items (flattened).
    Each invoice item becomes a separate row with repeated invoice header info.
    Enforces user-company access control for non-admin users.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Value mappings
    TTHAI_MAP = {
        1: "Hóa đơn mới",
        2: "Hóa đơn thay thế",
        3: "Hóa đơn điều chỉnh",
        4: "Hóa đơn bị thay thế",
        5: "Hóa đơn đã bị điều chỉnh",
        6: "Hóa đơn bị hủy",
    }

    KHMSHDON_MAP = {
        1: "Hóa đơn GTGT",
        2: "Hóa đơn bán hàng",
        3: "Phiếu xuất kho",
        4: "Hóa đơn khác",
    }

    TTXLY_MAP = {
        -1: "Chưa tra cứu",
        0: "Chưa kiểm tra",
        1: "Đã tiếp nhận",
        2: "Hóa đơn có sai sót",
        3: "Hóa đơn không đủ điều kiện",
        4: "Hóa đơn không hợp lệ",
        5: "Đã cấp mã hóa đơn",
        6: "Tổng cục thuế đã nhận",
        7: "Hóa đơn giả mạo",
        8: "Đã kiểm tra",
    }

    # For non-admin users, restrict access to assigned companies
    user_repo = UserRepository(conn)
    company_repo = CompanyRepository(conn)
    
    if current_user.role != "admin":
        # Get user's assigned companies
        user_companies = user_repo.get_user_companies(current_user.id)
        company_tax_codes = [comp['tax_code'] for comp in user_companies]
        
        if not company_tax_codes:
            # User has no assigned companies, return empty result
            raise HTTPException(status_code=403, detail="No companies assigned to user")
        
        # Build where clause with company restrictions
        where_clause, params = build_invoice_where_clause(from_date, to_date, tax_code, buyer_tax_code, search)
        
        # Add company restriction for non-admin users
        company_placeholders = ",".join(["%s"] * len(company_tax_codes))
        
        if where_clause == "1=1":
            # If no other filters, just add company filter
            where_clause = f"(i.nbmst IN ({company_placeholders}) OR i.nmmst IN ({company_placeholders}))"
            params = company_tax_codes + company_tax_codes
        else:
            # If there are other filters, add company restriction to existing where clause
            where_clause = f"({where_clause}) AND (i.nbmst IN ({company_placeholders}) OR i.nmmst IN ({company_placeholders}))"
            params = params + company_tax_codes + company_tax_codes
    else:
        # Admin users can export all invoices
        where_clause, params = build_invoice_where_clause(from_date, to_date, tax_code, buyer_tax_code, search)

    # Fetch flattened data with LEFT JOIN to include invoices without items
    # Note: ii.tthue is TEXT type, needs casting
    data_sql = f"""
        SELECT
            -- Invoice ID for tracking
            i.id AS invoice_id,
            -- Invoice header info (14 columns)
            i.khhdon, i.shdon, i.tdlap, i.dvtte, i.tgia,
            i.nbten, i.nbmst, i.nbdchi, i.nky, i.mhdon, i.ncma,
            i.nmten, i.nmmst, i.nmdchi,
            -- Invoice totals (3 columns)
            i.tgtcthue, i.tgtthue, i.tgtttbso,
            -- Invoice status (3 columns)
            i.khmshdon, i.tthai, i.ttxly,
            -- Item detail (14 columns)
            ii.stt AS item_stt, ii.tchat, ii.mhhdvu, ii.ten AS item_ten,
            ii.dvtinh, ii.sluong, ii.dgia, ii.tlckhau, ii.stckhau,
            ii.ltsuat, ii.tsuat, ii.thtien,
            CAST(NULLIF(ii.tthue, '') AS DOUBLE PRECISION) AS item_tthue,
            (COALESCE(ii.thtien, 0) + COALESCE(CAST(NULLIF(ii.tthue, '') AS DOUBLE PRECISION), 0)) AS thtcthue
        FROM invoices i
        LEFT JOIN invoice_items ii ON i.id = ii.idhdon
        WHERE {where_clause}
        ORDER BY i.nky DESC NULLS LAST, i.tdlap DESC NULLS LAST, i.shdon DESC NULLS LAST, ii.stt ASC NULLS LAST
    """
    
    with conn.cursor() as cur:
        cur.execute(data_sql, params if params else None)
        rows = cur.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Chi tiết hóa đơn"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    number_format = '#,##0'
    
    # Headers (33 columns)
    headers = [
        # Invoice header (15)
        "STT", "Ký hiệu mẫu số HĐ", "Ký hiệu HĐ", "Số HĐ", "Ngày lập", "ĐVT tiền", "Tỷ giá",
        "Tên NB", "MST NB", "Địa chỉ NB", "Ngày ký", "Mã HĐ", "Ngày cấp mã",
        "Tên NM", "MST NM", "Địa chỉ NM",
        # Invoice totals (3)
        "Tiền chưa thuế", "Tiền thuế", "Tổng tiền",
        # Invoice status (2)
        "Trạng thái", "Kết quả kiểm tra",
        # Item detail (14)
        "STT dòng", "Tính chất", "Mã HHDV", "Tên hàng hóa, dịch vụ",
        "ĐVT", "Số lượng", "Đơn giá", "TL chiết khấu", "ST chiết khấu",
        "Loại thuế suất", "Thuế suất", "Thành tiền", "Tiền thuế", "TT có thuế"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Data rows
    last_invoice_id = None
    for row_idx, row in enumerate(rows, 2):
        data = dict(row)
        invoice_id = data.get('invoice_id')
        is_first_item = (invoice_id != last_invoice_id)
        col = 1
        
        # STT
        ws.cell(row=row_idx, column=col, value=row_idx - 1).border = thin_border
        col += 1
        
        # Ký hiệu mẫu số HĐ (khmshdon - mapped)
        khmshdon_val = data.get('khmshdon')
        ws.cell(row=row_idx, column=col, value=KHMSHDON_MAP.get(khmshdon_val, str(khmshdon_val) if khmshdon_val else '')).border = thin_border
        col += 1
        
        # Invoice header columns
        ws.cell(row=row_idx, column=col, value=data.get('khhdon') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('shdon') or '').border = thin_border
        col += 1
        tdlap = data.get('tdlap')
        ws.cell(row=row_idx, column=col, value=tdlap[:10] if tdlap else '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('dvtte') or 'VND').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('tgia') or 1).border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('nbten') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('nbmst') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('nbdchi') or '').border = thin_border
        col += 1
        nky = data.get('nky')
        ws.cell(row=row_idx, column=col, value=nky[:10] if nky else '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('mhdon') or '').border = thin_border
        col += 1
        ncma = data.get('ncma')
        ws.cell(row=row_idx, column=col, value=ncma[:10] if ncma else '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('nmten') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('nmmst') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('nmdchi') or '').border = thin_border
        col += 1
        
        # Invoice totals (number format) - only on first item
        if is_first_item:
            cell = ws.cell(row=row_idx, column=col, value=float(data.get('tgtcthue') or 0))
            cell.border = thin_border
            cell.number_format = number_format
        else:
            ws.cell(row=row_idx, column=col, value='').border = thin_border
        col += 1
        if is_first_item:
            cell = ws.cell(row=row_idx, column=col, value=float(data.get('tgtthue') or 0))
            cell.border = thin_border
            cell.number_format = number_format
        else:
            ws.cell(row=row_idx, column=col, value='').border = thin_border
        col += 1
        if is_first_item:
            cell = ws.cell(row=row_idx, column=col, value=float(data.get('tgtttbso') or 0))
            cell.border = thin_border
            cell.number_format = number_format
        else:
            ws.cell(row=row_idx, column=col, value='').border = thin_border
        col += 1
        
        # Invoice status (mapped values) - Loại HĐ removed (now in header area)
        tthai_val = data.get('tthai')
        ws.cell(row=row_idx, column=col, value=TTHAI_MAP.get(tthai_val, str(tthai_val) if tthai_val else '')).border = thin_border
        col += 1
        ttxly_val = data.get('ttxly')
        ws.cell(row=row_idx, column=col, value=TTXLY_MAP.get(ttxly_val, str(ttxly_val) if ttxly_val else '')).border = thin_border
        col += 1
        
        # Item detail columns
        ws.cell(row=row_idx, column=col, value=data.get('item_stt') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('tchat') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('mhhdvu') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('item_ten') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('dvtinh') or '').border = thin_border
        col += 1
        cell = ws.cell(row=row_idx, column=col, value=float(data.get('sluong') or 0) if data.get('sluong') else '')
        cell.border = thin_border
        col += 1
        cell = ws.cell(row=row_idx, column=col, value=float(data.get('dgia') or 0) if data.get('dgia') else '')
        cell.border = thin_border
        cell.number_format = number_format
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('tlckhau') or '').border = thin_border
        col += 1
        cell = ws.cell(row=row_idx, column=col, value=float(data.get('stckhau') or 0) if data.get('stckhau') else '')
        cell.border = thin_border
        cell.number_format = number_format
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('ltsuat') or '').border = thin_border
        col += 1
        ws.cell(row=row_idx, column=col, value=data.get('tsuat') or '').border = thin_border
        col += 1
        cell = ws.cell(row=row_idx, column=col, value=float(data.get('thtien') or 0) if data.get('thtien') else '')
        cell.border = thin_border
        cell.number_format = number_format
        col += 1
        cell = ws.cell(row=row_idx, column=col, value=float(data.get('item_tthue') or 0) if data.get('item_tthue') else '')
        cell.border = thin_border
        cell.number_format = number_format
        col += 1
        cell = ws.cell(row=row_idx, column=col, value=float(data.get('thtcthue') or 0) if data.get('thtcthue') else '')
        cell.border = thin_border
        cell.number_format = number_format
        
        # Update last_invoice_id for next iteration
        last_invoice_id = invoice_id
    
    # Auto-size columns (with reasonable limits)
    column_widths = [
        5, 18, 12, 10, 12, 8, 8,  # STT, Ký hiệu mẫu số HĐ, Ký hiệu HĐ, Số HĐ, Ngày lập, ĐVT tiền, Tỷ giá
        25, 15, 30, 12, 15, 12,  # Tên/MST/ĐC NB, Ngày ký, Mã HĐ, Ngày cấp mã
        25, 15, 30,  # Tên/MST/ĐC NM
        15, 12, 15,  # Tiền chưa thuế, Tiền thuế, Tổng tiền
        22, 22,  # Trạng thái, Kết quả kiểm tra
        8, 10, 12, 35,  # STT dòng, Tính chất, Mã HHDV, Tên HH
        10, 10, 12, 12, 12,  # ĐVT, SL, Đơn giá, TL CK, ST CK
        15, 10, 12, 12, 12  # Loại TS, TS, Thành tiền, Tiền thuế, TT có thuế
    ]
    
    for idx, width in enumerate(column_widths, 1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date range
    filename = f"chitiet_hoadon_{from_date or 'all'}_{to_date or 'all'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.get("/stats", response_model=InvoiceStatsResponse)
async def get_stats(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    current_user: UserAuth = Depends(get_current_user),
    conn = Depends(get_db),
):
    """
    Get invoice statistics for dashboard.
    Enforces user-company access control for non-admin users.
    """
    # For non-admin users, restrict access to assigned companies
    user_repo = UserRepository(conn)
    company_repo = CompanyRepository(conn)
    
    # Build date filter using COALESCE(nky, tdlap):
    # - nky (ngày ký) is preferred; falls back to tdlap (ngày lập) when nky IS NULL
    conditions = []
    params = []

    if from_date:
        conditions.append("COALESCE(nky, tdlap) >= %s")
        params.append(from_date.isoformat())
    if to_date:
        conditions.append("COALESCE(nky, tdlap) <= %s")
        params.append(to_date.isoformat() + "T23:59:59")

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    # Apply company access control for non-admin users
    if current_user.role != "admin":
        # Get user's assigned companies
        user_companies = user_repo.get_user_companies(current_user.id)
        company_tax_codes = [comp['tax_code'] for comp in user_companies]
        
        if not company_tax_codes:
            # User has no assigned companies, return zero stats
            return InvoiceStatsResponse(
                total_invoices=0,
                total_purchase=0,
                total_sold=0,
                total_amount=0.0,
                total_tax=0.0,
                invoices_by_month={},
            )
        
        # Add company restriction to where clause
        company_placeholders = ",".join(["%s"] * len(company_tax_codes))
        
        if where_clause == "1=1":
            # If no date filters, just add company filter
            where_clause = f"(nbmst IN ({company_placeholders}) OR nmmst IN ({company_placeholders}))"
            params = company_tax_codes + company_tax_codes
        else:
            # If there are date filters, add company restriction
            where_clause = f"({where_clause}) AND (nbmst IN ({company_placeholders}) OR nmmst IN ({company_placeholders}))"
            params = params + company_tax_codes + company_tax_codes

    with conn.cursor() as cur:
        # Total counts
        cur.execute(f"""
            SELECT
                COUNT(*) as total,
                COALESCE(SUM(tgtcthue), 0) as total_amount,
                COALESCE(SUM(tgtthue), 0) as total_tax
            FROM invoices
            WHERE {where_clause}
        """, params if params else None)

        row = cur.fetchone()
        total_invoices = row['total']
        total_amount = float(row['total_amount']) if row['total_amount'] else 0.0
        total_tax = float(row['total_tax']) if row['total_tax'] else 0.0

        # Invoices by month (extract year-month from tdlap string)
        cur.execute(f"""
            SELECT
                SUBSTRING(tdlap FROM 1 FOR 7) as month,
                COUNT(*) as cnt
            FROM invoices
            WHERE tdlap IS NOT NULL AND {where_clause}
            GROUP BY month
            ORDER BY month DESC
            LIMIT 12
        """, params if params else None)

        by_month = {r['month']: r['cnt'] for r in cur.fetchall() if r['month']}

    return InvoiceStatsResponse(
        total_invoices=total_invoices,
        total_purchase=0,
        total_sold=0,
        total_amount=total_amount,
        total_tax=total_tax,
        invoices_by_month=by_month,
    )


@router.get("/{invoice_id}", response_model=InvoiceDetail)
async def get_invoice(
    invoice_id: str,
    current_user: UserAuth = Depends(get_current_user),
    conn = Depends(get_db),
):
    """
    Get single invoice with items.
    Enforces user-company access control for non-admin users.
    """
    # For non-admin users, check if they have access to this invoice's companies
    user_repo = UserRepository(conn)
    company_repo = CompanyRepository(conn)
    
    if current_user.role != "admin":
        # Get user's assigned companies
        user_companies = user_repo.get_user_companies(current_user.id)
        company_tax_codes = [comp['tax_code'] for comp in user_companies]
        
        if not company_tax_codes:
            # User has no assigned companies
            raise HTTPException(status_code=403, detail="No companies assigned to user")
        
        # Check if the invoice belongs to one of the user's companies
        with conn.cursor() as cur:
            # Get invoice header to check company tax codes
            cur.execute("""
                SELECT id, nbmst, nbten, nmmst, nmten, shdon, khhdon, khmshdon,
                       tdlap, tgtcthue, tgtthue, tgtttbso, tthai,
                       nbdchi, nmdchi, dvtte, tgtttbchu, htttoan
                FROM invoices
                WHERE id = %s
            """, (invoice_id,))
            
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Invoice not found")
            
            invoice_data = dict(row)
            
            # Check if user has access to either the seller or buyer company
            seller_tax_code = invoice_data.get('nbmst')
            buyer_tax_code = invoice_data.get('nmmst')
            
            has_access = (seller_tax_code in company_tax_codes) or (buyer_tax_code in company_tax_codes)
            
            if not has_access:
                raise HTTPException(status_code=403, detail="Access denied to this invoice")
        
        # If we reach here, user has access, so get the full invoice data
        with conn.cursor() as cur:
            # Get invoice header
            cur.execute("""
                SELECT id, nbmst, nbten, nmmst, nmten, shdon, khhdon, khmshdon,
                       tdlap, tgtcthue, tgtthue, tgtttbso, tthai,
                       nbdchi, nmdchi, dvtte, tgtttbchu, htttoan
                FROM invoices
                WHERE id = %s
            """, (invoice_id,))

            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Invoice not found")

            invoice_data = dict(row)

            # Get items
            cur.execute("""
                SELECT id, idhdon, stt, ten, dvtinh, sluong, dgia, thtien, tsuat
                FROM invoice_items
                WHERE idhdon = %s
                ORDER BY stt
            """, (invoice_id,))

            items = [InvoiceItemResponse(**dict(r)) for r in cur.fetchall()]

        invoice_data["items"] = items
        return InvoiceDetail(**invoice_data)
    else:
        # Admin users can access any invoice
        with conn.cursor() as cur:
            # Get invoice header
            cur.execute("""
                SELECT id, nbmst, nbten, nmmst, nmten, shdon, khhdon, khmshdon,
                       tdlap, tgtcthue, tgtthue, tgtttbso, tthai,
                       nbdchi, nmdchi, dvtte, tgtttbchu, htttoan
                FROM invoices
                WHERE id = %s
            """, (invoice_id,))

            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Invoice not found")

            invoice_data = dict(row)

            # Get items
            cur.execute("""
                SELECT id, idhdon, stt, ten, dvtinh, sluong, dgia, thtien, tsuat
                FROM invoice_items
                WHERE idhdon = %s
                ORDER BY stt
            """, (invoice_id,))

            items = [InvoiceItemResponse(**dict(r)) for r in cur.fetchall()]

        invoice_data["items"] = items
        return InvoiceDetail(**invoice_data)
